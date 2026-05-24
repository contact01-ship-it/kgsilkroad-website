"""
Re-parses all downloaded DSN price list Excel files into dsn_products.json.
Handles two layouts:
  A) "Importantly!" note in row 0, real headers in row 1
  B) Headers in row 0 (brand-specific formats)
"""
import json
import openpyxl
from pathlib import Path

PRICELISTS_DIR = Path("dsn_pricelists")

# Keyword → canonical field name
FIELD_KEYWORDS = {
    "product_name": [
        "description", "product", "name", "item name", "standard description",
        "product description", "new products",
    ],
    "sku": ["nutra #", "nutra#", "item #", "item number", "sku", "part #", "upc code", "item code", "prod #"],
    "upc": ["upc", "ean", "barcode", "gtin"],
    "size": ["size", "item size", "servings", "ounces", "oz", "volume"],
    "case_qty": [
        "case qty", "case\nqty", "tray/case size", "units per case", "case pack",
        "min. case qty", "cs qty", "case\npack", "cases/\ntl", "qty per case", "pcs/case",
    ],
    "price": [
        "price", "unit price", "price per unit", "whlsle", "international whsl",
        "logistic price", "wholesale", "msrp", "cost", "24-99 bottle price",
        "100+ bottle price", "price for pallet",
    ],
    "form": ["form"],
    "category": ["category", "main catalog category", "product category"],
    "rating": ["rating"],
    "weight_kg": ["weight, kg", "unit weight", "weight"],
}


def find_header_row(rows):
    """Return (header_idx, headers_list) — skips 'Importantly!' notes."""
    for i, row in enumerate(rows[:10]):
        cells = [str(c).strip() if c is not None else "" for c in row]
        # Skip rows that are empty or start with a long note
        if not any(cells):
            continue
        first = cells[0].lower()
        if first.startswith("importantly") or first.startswith("- the minimum"):
            continue
        # Need at least 3 non-empty cells to be a header row
        non_empty = [c for c in cells if c]
        if len(non_empty) >= 3:
            return i, cells
    return None, None


def match_header(header: str):
    """Match a header cell text to a canonical field name."""
    h = header.lower().strip().rstrip(":").replace("\n", " ")
    for field, keywords in FIELD_KEYWORDS.items():
        for kw in keywords:
            if kw in h or h in kw:
                return field
    return None


def parse_file(path: Path, brand_name: str, category: str) -> list:
    products = []
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        print(f"  ! Cannot open {path.name}: {e}")
        return products

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = [list(row) for row in ws.iter_rows(values_only=True)]
        if not rows:
            continue

        header_idx, raw_headers = find_header_row(rows)
        if header_idx is None:
            continue

        # Build column map: col_index → canonical field
        col_map = {}
        for i, h in enumerate(raw_headers):
            if not h:
                continue
            field = match_header(h)
            if field and field not in col_map.values():
                col_map[i] = field

        if not col_map:
            continue

        # Check if this sheet has useful product data
        has_name_or_sku = any(v in ("product_name", "sku") for v in col_map.values())
        if not has_name_or_sku:
            continue

        # Store all raw headers for extra_data
        for data_row in rows[header_idx + 1:]:
            cells = [str(c).strip() if c is not None else "" for c in data_row]

            # Skip empty rows and sub-headers
            non_empty = [c for c in cells if c]
            if len(non_empty) < 2:
                continue
            # Skip rows that look like repeated headers
            if cells[0].lower().strip() in ("nutra #", "item #", "upc", "sku", "description", "product"):
                continue

            row_data = {"brand": brand_name, "category": category}
            for col_idx, field in col_map.items():
                if col_idx < len(cells) and cells[col_idx]:
                    row_data[field] = cells[col_idx]

            # Must have a name or SKU to be useful
            if not row_data.get("product_name") and not row_data.get("sku"):
                continue

            products.append(row_data)

        # Only use first matching sheet to avoid duplicates
        if products:
            break

    wb.close()
    return products


def main():
    # Load brand metadata
    brands_file = Path("dsn_brands.json")
    brand_meta = {}
    if brands_file.exists():
        with open(brands_file) as f:
            brands = json.load(f)
        for b in brands:
            pl = b.get("pricelist_local", "")
            if pl:
                brand_meta[Path(pl).name] = {
                    "name": b["name"],
                    "category": b.get("category", ""),
                }

    all_products = []
    files = sorted(PRICELISTS_DIR.glob("*.xlsx")) + sorted(PRICELISTS_DIR.glob("*.xls"))

    for path in files:
        meta = brand_meta.get(path.name, {})
        brand_name = meta.get("name") or path.stem.replace("_", " ")
        category = meta.get("category", "")
        print(f"  Parsing [{category:2}] {path.name} ...", end=" ")
        rows = parse_file(path, brand_name, category)
        print(f"{len(rows)} rows")
        all_products.extend(rows)

    out = Path("dsn_products.json")
    with open(out, "w", encoding="utf-8") as f:
        json.dump(all_products, f, ensure_ascii=False, indent=2)

    print(f"\n✓ Saved {len(all_products)} product rows to {out}")

    # Quick stats
    with_price = sum(1 for p in all_products if p.get("price"))
    with_upc = sum(1 for p in all_products if p.get("upc"))
    with_size = sum(1 for p in all_products if p.get("size"))
    brands_covered = len(set(p["brand"] for p in all_products))
    print(f"  Brands covered: {brands_covered}")
    print(f"  With price:     {with_price}")
    print(f"  With UPC:       {with_upc}")
    print(f"  With size:      {with_size}")


if __name__ == "__main__":
    main()
