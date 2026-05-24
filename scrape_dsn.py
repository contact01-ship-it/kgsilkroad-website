import asyncio
import json
import os
import re
import httpx
from pathlib import Path
from playwright.async_api import async_playwright, Download

BASE_URL = "https://dsn.group"
EMAIL = "contact01@kgsilkroad.com"
PASSWORD = "Marleshdsn"
PRICELISTS_DIR = Path("dsn_pricelists")
IMAGES_DIR = Path("dsn_images")
PRICELISTS_DIR.mkdir(exist_ok=True)
IMAGES_DIR.mkdir(exist_ok=True)


async def wait_for_page(page, state="domcontentloaded"):
    try:
        await page.wait_for_load_state(state, timeout=15000)
    except Exception:
        pass  # continue even if timeout


async def goto(page, url):
    try:
        await page.goto(url, wait_until="domcontentloaded", timeout=60000)
    except Exception:
        pass
    await page.wait_for_timeout(1500)


async def dismiss_popups(page):
    for sel in [
        "button:has-text('Accept')", "button:has-text('I Accept')", "button:has-text('OK')",
        "button:has-text('Close')", "[class*='cookie'] button", "[id*='cookie'] button",
        "[class*='consent'] button", ".cc-btn", "#onetrust-accept-btn-handler"
    ]:
        try:
            btn = await page.query_selector(sel)
            if btn and await btn.is_visible():
                await btn.click()
                await page.wait_for_timeout(500)
                break
        except:
            pass


async def login(page):
    print("→ Opening dsn.group...")
    await goto(page, BASE_URL)
    await dismiss_popups(page)

    # Print all links on home page for debugging
    all_links = await page.eval_on_selector_all(
        "a[href]",
        "els => els.map(e => ({text: e.innerText.trim(), href: e.href, visible: e.offsetParent !== null})).filter(e => e.text && e.text.length < 30)"
    )
    print("  Home page links:")
    for lnk in all_links[:40]:
        print(f"    [{'+' if lnk.get('visible') else '-'}] {lnk.get('text','')[:30]:30} → {lnk.get('href','')}")

    # Try known login URLs first
    login_url = None
    for candidate in [
        f"{BASE_URL}/login", f"{BASE_URL}/account/login", f"{BASE_URL}/sign-in",
        f"{BASE_URL}/user/login", f"{BASE_URL}/auth/login", f"{BASE_URL}/wp-login.php",
        f"{BASE_URL}/my-account",
    ]:
        # Check if any anchor points there
        for lnk in all_links:
            if candidate in lnk.get("href", "") or lnk.get("href", "").rstrip("/") == candidate:
                login_url = lnk["href"]
                break
        if login_url:
            break

    if login_url:
        print(f"  → Found login URL: {login_url}")
        await goto(page, login_url)
    else:
        # Try clicking a visible login link
        for sel in [
            "a:has-text('Log in')", "a:has-text('Login')", "a:has-text('Sign in')",
            "a:has-text('Sign In')", "a[href*='login']:visible", "a[href*='sign-in']:visible"
        ]:
            try:
                btn = await page.query_selector(sel)
                if btn and await btn.is_visible():
                    await btn.click()
                    await wait_for_page(page)
                    break
            except:
                pass

    await dismiss_popups(page)
    await page.wait_for_timeout(2000)
    await dismiss_popups(page)

    print(f"  Current URL: {page.url}")
    print(f"  Page title: {await page.title()}")

    # Use JS to fill the form (bypasses visibility restrictions)
    filled = await page.evaluate(f"""() => {{
        const inputs = document.querySelectorAll('input');
        let emailFilled = false, passwordFilled = false;
        for (const inp of inputs) {{
            const t = inp.type.toLowerCase();
            const n = (inp.name || '').toLowerCase();
            const p = (inp.placeholder || '').toLowerCase();
            const id = (inp.id || '').toLowerCase();
            if (!emailFilled && (t === 'email' || n.includes('email') || n.includes('user') || p.includes('email') || p.includes('user') || id.includes('email') || id.includes('user'))) {{
                inp.value = '{EMAIL}';
                inp.dispatchEvent(new Event('input', {{bubbles: true}}));
                inp.dispatchEvent(new Event('change', {{bubbles: true}}));
                emailFilled = true;
            }} else if (!passwordFilled && t === 'password') {{
                inp.value = 'Marleshdsn';
                inp.dispatchEvent(new Event('input', {{bubbles: true}}));
                inp.dispatchEvent(new Event('change', {{bubbles: true}}));
                passwordFilled = true;
            }}
        }}
        return {{emailFilled, passwordFilled}};
    }}""")
    print(f"  Form fill result: {filled}")

    if not filled.get("emailFilled"):
        print("  ! Could not fill form via JS, dumping body HTML...")
        html = await page.inner_html("body")
        print(html[:3000])
        return

    await page.wait_for_timeout(500)

    # Submit the form
    submitted = await page.evaluate("""() => {
        const btn = document.querySelector('button[type="submit"], input[type="submit"], button.login, button.signin');
        if (btn) { btn.click(); return true; }
        const form = document.querySelector('form');
        if (form) { form.submit(); return true; }
        return false;
    }""")
    print(f"  Form submitted: {submitted}")

    await wait_for_page(page)
    await page.wait_for_timeout(2000)
    print(f"✓ After login — URL: {page.url}")


async def find_brands_section(page, section_name: str) -> list:
    """Navigate to US/EU Brands section and collect brand info."""
    print(f"\n→ Looking for '{section_name}' section...")
    brands = []

    # Try nav/menu links
    selectors_to_try = [
        f"a:has-text('{section_name}')",
        f"nav a:has-text('{section_name.split()[0]}')",
        f"[href*='us-brand'], [href*='eu-brand'], [href*='brand']",
    ]

    link = None
    for sel in selectors_to_try:
        try:
            link = await page.query_selector(sel)
            if link:
                text = await link.inner_text()
                if section_name.split()[0].lower() in text.lower():
                    break
                link = None
        except:
            pass

    # Known URL mapping from site exploration
    known_urls = {
        "US Brands": f"{BASE_URL}/brands/",
        "EU Brands": f"{BASE_URL}/profilnie-brendi/",
    }
    if section_name in known_urls:
        await goto(page, known_urls[section_name])
        print(f"  Navigated to: {page.url}")
    elif link:
        await link.click()
        await wait_for_page(page)
        await page.wait_for_timeout(1500)
        print(f"  Navigated to: {page.url}")
    else:
        print(f"  ! Could not find '{section_name}' link, trying URL patterns...")
        slug = section_name.lower().replace(" ", "-")
        for url in [f"{BASE_URL}/{slug}", f"{BASE_URL}/brands/{slug}", f"{BASE_URL}/{slug}s"]:
            await goto(page, url)
            if page.url == url:
                break

    # Scroll to load all brands
    prev_height = 0
    for _ in range(10):
        height = await page.evaluate("document.body.scrollHeight")
        if height == prev_height:
            break
        await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
        await page.wait_for_timeout(800)
        prev_height = height

    # Collect brand cards/links
    brand_selectors = [
        ".brand-card", ".brand-item", ".brand", "[class*='brand']",
        ".card", ".item", "article", ".product-item"
    ]

    collected = set()
    for sel in brand_selectors:
        items = await page.query_selector_all(sel)
        if len(items) > 3:
            for item in items:
                brand = {}
                # Name
                for name_sel in ["h1", "h2", "h3", "h4", ".title", ".name", "[class*='name']", "[class*='title']"]:
                    try:
                        el = await item.query_selector(name_sel)
                        if el:
                            text = (await el.inner_text()).strip()
                            if text and len(text) > 1:
                                brand["name"] = text
                                break
                    except:
                        pass

                # Link
                for link_sel in ["a[href]", "a"]:
                    try:
                        el = await item.query_selector(link_sel)
                        if el:
                            href = await el.get_attribute("href")
                            if href and href not in ("#", ""):
                                if href.startswith("/"):
                                    href = BASE_URL + href
                                brand["url"] = href
                                break
                    except:
                        pass

                # Image
                try:
                    img = await item.query_selector("img")
                    if img:
                        brand["image_url"] = await img.get_attribute("src") or await img.get_attribute("data-src") or ""
                except:
                    brand["image_url"] = ""

                if brand.get("name") and brand.get("url") and brand["url"] not in collected:
                    collected.add(brand["url"])
                    brands.append(brand)
            if brands:
                break

    # Fallback: collect all links that look like brand pages
    if not brands:
        all_links = await page.eval_on_selector_all(
            "a[href]",
            "els => els.map(e => ({href: e.href, text: e.innerText.trim()}))"
        )
        for item in all_links:
            href = item.get("href", "")
            text = item.get("text", "").strip()
            if (BASE_URL in href and text and len(text) > 2 and
                    any(w in href.lower() for w in ["brand", "supplier", "vendor", "catalog"]) and
                    href not in collected):
                collected.add(href)
                brands.append({"name": text, "url": href, "image_url": ""})

    print(f"  Found {len(brands)} brands in '{section_name}'")
    return brands


async def download_file(url: str, dest: Path, cookies: dict = None) -> bool:
    if not url:
        return False
    try:
        async with httpx.AsyncClient(timeout=60, follow_redirects=True, cookies=cookies) as client:
            resp = await client.get(url)
            if resp.status_code == 200:
                dest.write_bytes(resp.content)
                return True
    except Exception as e:
        print(f"    ! Download error: {e}")
    return False


def sanitize(name: str) -> str:
    return re.sub(r'[^\w\-_.]', '_', name)[:80]


async def process_brand(page, brand: dict, category: str, index: int, total: int, cookies: dict) -> dict:
    name = brand.get("name", f"brand_{index}")
    url = brand.get("url", "")
    print(f"\n[{index}/{total}] {category} | {name}")

    result = {
        "name": name,
        "category": category,
        "url": url,
        "description": "",
        "image_local": "",
        "pricelist_local": "",
    }

    if not url:
        return result

    try:
        await goto(page, url)

        # Description
        for sel in ["[class*='description']", "[class*='about']", "p", ".content"]:
            try:
                el = await page.query_selector(sel)
                if el:
                    text = (await el.inner_text()).strip()
                    if len(text) > 20:
                        result["description"] = text[:500]
                        break
            except:
                pass

        # Brand logo/image — use image from listing or find on page
        img_url = brand.get("image_url", "")
        if not img_url:
            for sel in ["[class*='logo'] img", "[class*='brand'] img", "header img", ".hero img", "img"]:
                try:
                    img = await page.query_selector(sel)
                    if img:
                        img_url = await img.get_attribute("src") or ""
                        if img_url:
                            break
                except:
                    pass

        if img_url:
            if img_url.startswith("/"):
                img_url = BASE_URL + img_url
            ext = img_url.split("?")[0].rsplit(".", 1)[-1].lower() if "." in img_url else "jpg"
            if ext not in ("png", "jpg", "jpeg", "webp", "svg", "gif"):
                ext = "jpg"
            img_file = IMAGES_DIR / f"{sanitize(name)}.{ext}"
            ok = await download_file(img_url, img_file, cookies)
            if ok:
                result["image_local"] = str(img_file)
                print(f"  ✓ Image: {img_file.name}")

        # Price list button — look for download link
        pricelist_url = ""
        for sel in [
            "a:has-text('PRICE LIST')", "a:has-text('Price List')", "a:has-text('Pricelist')",
            "a[href*='price']", "a[href*='excel']", "a[href*='.xlsx']", "a[href*='.xls']",
            "button:has-text('PRICE')", "[class*='price-list'] a", "[class*='pricelist'] a"
        ]:
            try:
                el = await page.query_selector(sel)
                if el:
                    pricelist_url = await el.get_attribute("href") or ""
                    if pricelist_url:
                        break
            except:
                pass

        if pricelist_url:
            if pricelist_url.startswith("/"):
                pricelist_url = BASE_URL + pricelist_url
            ext = pricelist_url.split("?")[0].rsplit(".", 1)[-1].lower()
            if ext not in ("xlsx", "xls", "csv"):
                ext = "xlsx"
            pl_file = PRICELISTS_DIR / f"{sanitize(name)}.{ext}"

            # Use Playwright download handler for protected files
            try:
                async with page.expect_download(timeout=30000) as dl_info:
                    btn = await page.query_selector(
                        "a:has-text('PRICE LIST'), a:has-text('Price List'), "
                        "a[href*='price'], a[href*='.xlsx'], a[href*='.xls']"
                    )
                    if btn:
                        await btn.click()
                download = await dl_info.value
                await download.save_as(pl_file)
                result["pricelist_local"] = str(pl_file)
                print(f"  ✓ Pricelist: {pl_file.name}")
            except Exception:
                # Fallback: direct download with session cookies
                ok = await download_file(pricelist_url, pl_file, cookies)
                if ok:
                    result["pricelist_local"] = str(pl_file)
                    print(f"  ✓ Pricelist (direct): {pl_file.name}")
                else:
                    print(f"  ! Pricelist download failed: {pricelist_url}")
        else:
            print(f"  - No price list found")

    except Exception as e:
        print(f"  ! Error: {e}")

    return result


def parse_excel_files(brands: list) -> list:
    try:
        import openpyxl
    except ImportError:
        print("! openpyxl not installed, skipping Excel parsing")
        return []

    all_products = []
    for brand in brands:
        pl = brand.get("pricelist_local", "")
        if not pl or not Path(pl).exists():
            continue

        brand_name = brand.get("name", "")
        print(f"  Parsing: {pl}")

        try:
            wb = openpyxl.load_workbook(pl, read_only=True, data_only=True)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    continue

                # Find header row (first non-empty row)
                header_row = None
                data_start = 0
                for i, row in enumerate(rows[:10]):
                    non_empty = [str(c).strip() for c in row if c is not None and str(c).strip()]
                    if len(non_empty) >= 2:
                        header_row = [str(c).strip() if c is not None else "" for c in row]
                        data_start = i + 1
                        break

                if not header_row:
                    continue

                # Map headers to standard fields
                col_map = {}
                header_lower = [h.lower() for h in header_row]
                for field, keywords in {
                    "product_name": ["product", "name", "description", "item", "sku"],
                    "size": ["size", "pack", "volume", "weight", "unit", "qty per"],
                    "price": ["price", "cost", "msrp", "wholesale", "usd", "$"],
                    "upc": ["upc", "barcode", "ean"],
                    "case_qty": ["case", "cs qty", "pcs", "per case", "case qty"],
                }.items():
                    for kw in keywords:
                        matches = [i for i, h in enumerate(header_lower) if kw in h]
                        if matches and field not in col_map:
                            col_map[field] = matches[0]

                for row in rows[data_start:]:
                    row_vals = [str(c).strip() if c is not None else "" for c in row]
                    if not any(row_vals):
                        continue

                    product = {"brand": brand_name, "category": brand.get("category", "")}
                    for field, idx in col_map.items():
                        if idx < len(row_vals):
                            product[field] = row_vals[idx]

                    # Fallback: use column indices if no mapping
                    if "product_name" not in product and row_vals:
                        product["product_name"] = row_vals[0]

                    if product.get("product_name") and product["product_name"] not in ("", "None"):
                        all_products.append(product)

            wb.close()
        except Exception as e:
            print(f"    ! Excel parse error for {pl}: {e}")

    return all_products


async def main():
    output_file = Path("dsn_brands.json")
    products_file = Path("dsn_products.json")

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False)  # visible to handle any JS challenges
        context = await browser.new_context(
            viewport={"width": 1440, "height": 900},
            accept_downloads=True,
        )
        page = await context.new_page()

        await login(page)

        # Debug: show page structure
        print(f"\n→ Current URL after login: {page.url}")
        nav_links = await page.eval_on_selector_all(
            "nav a, header a, [class*='nav'] a, [class*='menu'] a",
            "els => els.map(e => ({text: e.innerText.trim(), href: e.href})).filter(e => e.text)"
        )
        print("  Nav links found:")
        for link in nav_links[:30]:
            print(f"    {link['text'][:40]:40} → {link['href']}")

        # Get session cookies for direct downloads
        cookies_list = await context.cookies()
        cookies = {c["name"]: c["value"] for c in cookies_list}

        # Collect brands from US and EU sections
        all_brands = []
        us_brands = await find_brands_section(page, "US Brands")
        for b in us_brands:
            b["category"] = "US"
        all_brands.extend(us_brands)

        eu_brands = await find_brands_section(page, "EU Brands")
        for b in eu_brands:
            b["category"] = "EU"
        all_brands.extend(eu_brands)

        print(f"\n→ Total brands to process: {len(all_brands)}")

        # Process each brand
        results = []
        total = len(all_brands)
        for i, brand in enumerate(all_brands, 1):
            # Refresh cookies periodically
            if i % 10 == 0:
                cookies_list = await context.cookies()
                cookies = {c["name"]: c["value"] for c in cookies_list}

            result = await process_brand(page, brand, brand.get("category", ""), i, total, cookies)
            results.append(result)

            # Save progress
            with open(output_file, "w", encoding="utf-8") as f:
                json.dump(results, f, ensure_ascii=False, indent=2)

        await browser.close()

    # Parse all Excel files
    print("\n→ Parsing Excel price lists...")
    all_products = parse_excel_files(results)
    print(f"  Extracted {len(all_products)} product rows")

    with open(products_file, "w", encoding="utf-8") as f:
        json.dump(all_products, f, ensure_ascii=False, indent=2)

    print(f"\n✓ Done!")
    print(f"  Brands:   {output_file} ({len(results)} brands)")
    print(f"  Products: {products_file} ({len(all_products)} rows)")
    print(f"  Images:   {IMAGES_DIR}/ ({len(list(IMAGES_DIR.iterdir()))} files)")
    print(f"  Pricelists: {PRICELISTS_DIR}/ ({len(list(PRICELISTS_DIR.iterdir()))} files)")


if __name__ == "__main__":
    asyncio.run(main())
